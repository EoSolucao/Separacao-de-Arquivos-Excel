import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image
import flet as ft
from tkinter import filedialog, messagebox, Tk
import win32com.client
import pythoncom
import time
import random
from PIL import Image as PILImage
import tempfile
import traceback

# Constantes para a tabela dinâmica
xlRowField = 1
xlDataField = 2
xlColumnField = 3
xlPageField = 4
xlSum = -4105  # Constante do Excel para soma
xlCount = -4112  # Constante do Excel para contagem


def aplicar_formato(ws):
    fill_header = PatternFill(
        start_color="00008B", end_color="00008B", fill_type="solid"
    )
    font_header = Font(color="FFFFFF", bold=True)

    for cell in ws[4]:
        cell.fill = fill_header
        cell.font = font_header

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            if row[0].row % 2 == 0:
                cell.fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                )

    ws.sheet_view.showGridLines = False

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def wait_for_excel_availability(file_path, max_attempts=10, delay=2):
    for attempt in range(max_attempts):
        try:
            with open(file_path, "r+"):
                return True
        except IOError:
            time.sleep(delay)
    return False


def criar_tabela_dinamica(
    arquivo, linhas, colunas, valores, filtros, coluna_logo, coluna_titulo
):
    pythoncom.CoInitialize()
    excel = None
    wb = None

    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        if not wait_for_excel_availability(arquivo):
            print(f"Não foi possível acessar o arquivo: {arquivo}")
            return

        wb = excel.Workbooks.Open(arquivo)
        ws_info = wb.Worksheets("Dados")

        # Find the last row and column
        last_row = ws_info.Cells(ws_info.Rows.Count, "A").End(-4162).Row
        last_column = ws_info.Cells(4, ws_info.Columns.Count).End(-4159).Column
        print(f"Última linha: {last_row}, Última coluna: {last_column}")

        data_range = ws_info.Range(
            ws_info.Cells(4, 1), ws_info.Cells(last_row, last_column)
        )

        ws_pivot = wb.Worksheets.Add()
        ws_pivot.Name = "Informação"

        excel.ActiveWindow.DisplayGridlines = False

        # Add logo to the 'Informação' sheet
        if coluna_logo and os.path.isfile(coluna_logo):
            try:
                img = PILImage.open(coluna_logo)
                img_width, img_height = img.size
                dpi = 96
                max_width = max_height = 1.2 * dpi

                if img_width > max_width or img_height > max_height:
                    img.thumbnail((max_width, max_height))

                temp_logo_path = os.path.join(tempfile.gettempdir(), "temp_logo.png")
                img.save(temp_logo_path)

                ws_pivot.Shapes.AddPicture(
                    temp_logo_path,
                    LinkToFile=False,
                    SaveWithDocument=True,
                    Left=0,
                    Top=0,
                    Width=img.width,
                    Height=img.height,
                )
            except Exception as e:
                print(f"Erro ao adicionar logo na aba 'Informação': {e}")

        # Add title to the 'Informação' sheet
        ws_pivot.Cells(1, 3).Value = coluna_titulo
        ws_pivot.Cells(1, 3).Font.Bold = True
        ws_pivot.Cells(1, 3).Font.Size = 18

        pivot_cache = wb.PivotCaches().Create(SourceType=1, SourceData=data_range)
        pivot_table = pivot_cache.CreatePivotTable(
            TableDestination=ws_pivot.Range("A4"), TableName="TabelaDinamica"
        )

        # Print available fields
        print("Campos disponíveis na tabela dinâmica:")
        for field in pivot_table.PivotFields():
            print(f"- {field.Name}")

        def add_field(field_name, orientation):
            try:
                if field_name in [field.Name for field in pivot_table.PivotFields()]:
                    field = pivot_table.PivotFields(field_name)
                    field.Orientation = orientation
                    print(f"Campo '{field_name}' adicionado com sucesso.")
                    return True
                else:
                    print(f"Campo '{field_name}' não encontrado na tabela dinâmica.")
                    return False
            except Exception as e:
                print(f"Erro ao adicionar campo '{field_name}': {str(e)}")
                return False

        for campo in linhas.split(","):
            if campo.strip():
                add_field(campo.strip(), xlRowField)

        for campo in colunas.split(","):
            if campo.strip():
                add_field(campo.strip(), xlColumnField)

        for campo in valores.split(","):
            if campo.strip():
                if add_field(campo.strip(), xlDataField):
                    try:
                        value_field = pivot_table.PivotFields(campo.strip())
                        value_field.Function = xlSum
                    except:
                        try:
                            value_field.Function = xlCount
                        except Exception as e:
                            print(
                                f"Não foi possível definir uma função para o campo: {campo}. Erro: {str(e)}"
                            )

        for campo in filtros.split(","):
            if campo.strip():
                add_field(campo.strip(), xlPageField)

        pivot_table.RefreshTable()
        pivot_table.RowAxisLayout(1)

        ws_pivot.Columns("B").NumberFormat = "[hh]:mm"
        ws_pivot.Columns("C").NumberFormat = "[hh]:mm"
        ws_pivot.Columns("D").NumberFormat = "[hh]:mm"
        ws_pivot.Columns.AutoFit()

        wb.Save()
        print("Tabela dinâmica criada com sucesso.")
    except Exception as e:
        print(f"Erro ao criar tabela dinâmica: {str(e)}")
        print(f"Erro detalhado:\n{traceback.format_exc()}")
    finally:
        if wb:
            wb.Close(SaveChanges=True)
        if excel:
            excel.Quit()
        pythoncom.CoUninitialize()


def separar_e_ordenar_arquivos(
    arquivo_origem,
    pasta_destino,
    coluna_agrupamento,
    coluna_separacao,
    linhas,
    colunas,
    valores,
    filtros,
    coluna_logo,
    coluna_titulo,
):
    if not arquivo_origem or not pasta_destino:
        messagebox.showwarning(
            "Aviso", "Os campos de origem e resultado devem ser preenchidos!"
        )
        return

    messagebox.showinfo("Aguarde", "O código está em execução, por favor aguarde...")

    try:
        df_origem = pd.read_excel(arquivo_origem)
        print(f"Colunas no arquivo de origem: {', '.join(df_origem.columns)}")

        colunas_necessarias = (
            [coluna_agrupamento]
            + ([coluna_separacao] if coluna_separacao else [])
            + linhas.split(",")
            + colunas.split(",")
            + valores.split(",")
            + filtros.split(",")
        )
        colunas_necessarias = [
            col.strip() for col in colunas_necessarias if col.strip()
        ]
        colunas_faltantes = [
            col for col in colunas_necessarias if col not in df_origem.columns
        ]

        if colunas_faltantes:
            messagebox.showerror(
                "Erro",
                f"As seguintes colunas não foram encontradas no arquivo: {', '.join(colunas_faltantes)}",
            )
            return

        grupos = df_origem.groupby(coluna_agrupamento)

        if not os.path.exists(pasta_destino):
            os.makedirs(pasta_destino)

        with tempfile.TemporaryDirectory() as temp_dir:
            for grupo_nome, grupo_df in grupos:
                nome_arquivo = f"grupo_{grupo_nome}".replace("/", "_").replace(
                    "\\", "_"
                )
                caminho_arquivo_grupo = os.path.join(
                    pasta_destino, f"{nome_arquivo[:248]}.xlsx"
                )
                with pd.ExcelWriter(caminho_arquivo_grupo, engine="openpyxl") as writer:
                    grupo_df.to_excel(
                        writer, sheet_name="Dados", index=False, startrow=3
                    )

                    workbook = writer.book
                    worksheet = writer.sheets["Dados"]

                    if coluna_logo and os.path.isfile(coluna_logo):
                        try:
                            img = PILImage.open(coluna_logo)
                            img_width, img_height = img.size
                            dpi = 96
                            max_width = max_height = 2 * dpi

                            if img_width > max_width or img_height > max_height:
                                img.thumbnail((max_width, max_height))

                            temp_logo_path = os.path.join(temp_dir, "temp_logo.png")
                            img.save(temp_logo_path)

                            img = Image(temp_logo_path)
                            worksheet.add_image(img, "A1")
                        except Exception as e:
                            print(f"Erro ao adicionar logo: {e}")
                            messagebox.showwarning(
                                "Aviso", f"Não foi possível adicionar o logo: {e}"
                            )
                    else:
                        print(
                            f"Arquivo de logo não encontrado ou inválido: {coluna_logo}"
                        )

                    worksheet["C1"] = coluna_titulo
                    worksheet["C1"].font = Font(bold=True, size=18)
                    worksheet["C2"] = os.path.basename(caminho_arquivo_grupo)

                    aplicar_formato(worksheet)

                    if coluna_separacao:
                        for valor_separacao in grupo_df[coluna_separacao].unique():
                            if pd.notna(valor_separacao):
                                nome_aba = (
                                    str(valor_separacao)
                                    .replace("/", "_")
                                    .replace("\\", "_")[:31]
                                )
                                df_separado = grupo_df[
                                    grupo_df[coluna_separacao] == valor_separacao
                                ]
                                df_separado.to_excel(
                                    writer, sheet_name=nome_aba, index=False, startrow=3
                                )

                                worksheet_separado = writer.sheets[nome_aba]

                                if coluna_logo and os.path.isfile(coluna_logo):
                                    img = Image(temp_logo_path)
                                    worksheet_separado.add_image(img, "A1")

                                worksheet_separado["C1"] = coluna_titulo
                                worksheet_separado["C1"].font = Font(bold=True, size=18)
                                worksheet_separado["C2"] = nome_aba

                                aplicar_formato(worksheet_separado)

            for file_name in os.listdir(pasta_destino):
                if file_name.endswith(".xlsx"):
                    caminho_arquivo = os.path.join(pasta_destino, file_name)
                    print(f"Criando tabela dinâmica para o arquivo: {caminho_arquivo}")
                    criar_tabela_dinamica(
                        caminho_arquivo,
                        linhas,
                        colunas,
                        valores,
                        filtros,
                        coluna_logo,
                        coluna_titulo,
                    )

                    wb = load_workbook(caminho_arquivo)
                    ws_dados = wb["Dados"]

                    for row in range(5, ws_dados.max_row + 1):
                        cell = ws_dados.cell(row=row, column=11)
                        cell.number_format = "[hh]:mm"

                    for row in range(5, ws_dados.max_row + 1):
                        cell = ws_dados.cell(row=row, column=6)
                        cell.number_format = "dd/mm/yyyy"

                    wb.save(caminho_arquivo)

        messagebox.showinfo("Sucesso", "Arquivos separados e ordenados com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
        print(f"Erro detalhado:\n{traceback.format_exc()}")


def browse_file(target_text_field):
    root = Tk()
    root.withdraw()
    file_selected = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
    if file_selected:
        target_text_field.value = file_selected
        target_text_field.update()
    root.destroy()


def browse_folder(target_text_field):
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        target_text_field.value = folder_selected
        target_text_field.update()
    root.destroy()


def browse_logo(target_text_field):
    root = Tk()
    root.withdraw()
    file_selected = filedialog.askopenfilename(filetypes=[("Arquivos PNG", "*.png")])
    if file_selected:
        target_text_field.value = file_selected
        target_text_field.update()
    root.destroy()


def enviar_email_com_retry(outlook, mail_item, max_retries=3):
    for attempt in range(max_retries):
        try:
            mail_item.Send()
            return True
        except Exception as e:
            print(f"Tentativa {attempt + 1} falhou: {str(e)}")
            if attempt < max_retries - 1:
                time.sleep(random.uniform(5, 10))
            else:
                print(f"Falha ao enviar e-mail após {max_retries} tentativas.")
                return False


def enviar_emails(e, assunto, arquivo_envio, pasta_destino):
    try:
        df = pd.read_excel(arquivo_envio, sheet_name="Envio")
        outlook = win32com.client.Dispatch("outlook.application")

        for index, row in df.iterrows():
            sucesso = False
            tentativas = 0

            while not sucesso and tentativas < 1:
                try:
                    mail_item = outlook.CreateItem(0)
                    mail_item.Subject = f"{assunto} - {row['Arquivo']}"
                    mail_item.Body = row["Dados"]
                    mail_item.To = row["Email"]

                    arquivo_anexo = row["Arquivo"]
                    caminho_anexo = os.path.join(pasta_destino, arquivo_anexo)
                    if os.path.isfile(caminho_anexo):
                        mail_item.Attachments.Add(caminho_anexo)
                    else:
                        print(f"Arquivo não encontrado: {caminho_anexo}")
                        break

                    if enviar_email_com_retry(outlook, mail_item):
                        print(
                            f"E-mail enviado para {row['Email']} com o assunto '{mail_item.Subject}'."
                        )
                        messagebox.showinfo(
                            "Sucesso", f"Email enviado para {row['Email']}!"
                        )
                        sucesso = True
                    else:
                        raise Exception(
                            "Falha ao enviar e-mail após várias tentativas."
                        )

                    time.sleep(random.uniform(8, 12))

                except Exception as e:
                    tentativas += 1
                    if "movido ou excluído" in str(e) or "2147221238" in str(e):
                        print(
                            f"Erro ao enviar e-mail para {row['Email']}: O item foi movido ou excluído."
                        )
                        print(f"Tentando novamente ({tentativas}/1)...")
                        time.sleep(random.uniform(5, 10))
                    else:
                        print(f"Erro ao preparar e-mail para {row['Email']}: {str(e)}")
                        messagebox.showerror(
                            "Erro",
                            f"Erro ao preparar e-mail para {row['Email']}: {str(e)}",
                        )
                        break

        outlook = None
        messagebox.showinfo("Sucesso", "Processo de envio de e-mails concluído.")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao processar os e-mails: {e}")
        print(f"Erro detalhado:\n{traceback.format_exc()}")


def main(page: ft.Page):
    page.window_max_width = 380
    page.window_width = 380
    page.window_max_height = 1050
    page.window_height = 1050
    page.padding = 0
    page.title = "Separar, Ordenar e Enviar E-mails"

    title_text = ft.Text("Separar, Ordenar e Enviar E-mails", size=20, weight="bold")
    subtitle_text = ft.Text(
        "Dados para criar Tabela Dinâmica", size=11, color="gray", weight="bold"
    )
    subtitle_text2 = ft.Text(
        "Assunto para Enviar Email", size=11, color="gray", weight="bold"
    )
    subtitle_text3 = ft.Text("Título Planilha", size=11, color="gray", weight="bold")

    arquivo_input = ft.TextField(label="Arquivo de Origem", width=250, disabled=True)
    arquivo_btn = ft.IconButton(
        icon=ft.icons.SEARCH, on_click=lambda _: browse_file(arquivo_input)
    )

    pasta_input = ft.TextField(label="Pasta de Destino", width=250, disabled=True)
    pasta_btn = ft.IconButton(
        icon=ft.icons.SEARCH, on_click=lambda _: browse_folder(pasta_input)
    )

    arquivo_envio_input = ft.TextField(
        label="Planilha Envio Email", width=250, disabled=True
    )
    arquivo_envio_btn = ft.IconButton(
        icon=ft.icons.SEARCH, on_click=lambda _: browse_file(arquivo_envio_input)
    )

    coluna_logo = ft.TextField(label="Logo", width=250, disabled=True)
    coluna_logo_btn = ft.IconButton(
        icon=ft.icons.SEARCH, on_click=lambda _: browse_logo(coluna_logo)
    )

    coluna_titulo = ft.TextField(label="Título Planilha", width=250)
    coluna_agrupamento_input = ft.TextField(label="Coluna de Agrupamento", width=250)
    coluna_separacao_input = ft.TextField(label="Coluna de Separação", width=250)
    linhas_input = ft.TextField(label="Linhas", width=250)
    colunas_input = ft.TextField(label="Filtros", width=250)
    valores_input = ft.TextField(label="Colunas", width=250)
    filtros_input = ft.TextField(label="Valores", width=250)
    assunto_input = ft.TextField(label="Assunto", width=250)

    executar_btn = ft.ElevatedButton(
        text="Separar e Ordenar",
        on_click=lambda _: separar_e_ordenar_arquivos(
            arquivo_input.value,
            pasta_input.value,
            coluna_agrupamento_input.value,
            coluna_separacao_input.value,
            linhas_input.value,
            colunas_input.value,
            valores_input.value,
            filtros_input.value,
            coluna_logo.value,
            coluna_titulo.value,
        ),
    )

    enviar_email_btn = ft.ElevatedButton(
        text="Enviar Email",
        on_click=lambda _: enviar_emails(
            _, assunto_input.value, arquivo_envio_input.value, pasta_input.value
        ),
    )

    body = ft.Container(
        ft.Column(
            [
                title_text,
                ft.Row([arquivo_input, arquivo_btn], alignment="center"),
                ft.Row([pasta_input, pasta_btn], alignment="center"),
                ft.Row([arquivo_envio_input, arquivo_envio_btn], alignment="center"),
                ft.Row([coluna_logo, coluna_logo_btn], alignment="center"),
                coluna_agrupamento_input,
                coluna_separacao_input,
                subtitle_text3,
                coluna_titulo,
                subtitle_text,
                linhas_input,
                colunas_input,
                valores_input,
                filtros_input,
                subtitle_text2,
                assunto_input,
                executar_btn,
                enviar_email_btn,
            ],
            alignment="center",
            horizontal_alignment="center",
        ),
        gradient=ft.LinearGradient(
            begin=ft.alignment.top_left,
            end=ft.alignment.bottom_right,
            colors=["white", "#67b19f"],
        ),
        width=380,
        height=970,
        alignment=ft.alignment.center,
    )

    page.add(body)

    footer = ft.Text("EO SOLUÇÕES & CRIAÇÕES", size=12, color="gray")
    page.add(ft.Row([footer], alignment="center"))


ft.app(target=main)
